# хороший тон передавать зависимости: requirements.txt
from openpyxl import Workbook, load_workbook

# новый эксел-книга
wb = Workbook()

# берём активный лист (имени ещё нет, но который первый)
ws = wb.active
ws.title = 'табло'

wb.save('../docs/Report.xlsx')

# запись данных в существующий файл
wb = load_workbook('../docs/Report.xlsx')
ws = wb.active

# способ записи
ws['F1'] = 'привет' # мб также именем ячейки, буде определена
ws.cell(row=3, column=3, value='♥ любимые ♥')

wb.save('../docs/Report-привет.xlsx')

wb = Workbook()
ws = wb.active
# как забьём данные?
# заголовки
ws['A1'] = "ФИО"
ws['B1'] = 'Должность'
ws['C1'] = 'Отделъ'

employees = [
    ["Иванов И.И.", "Менеджер", "Отдел втюхивания"],
    ["Гарклидзе С.Т.", "Художник", "Отдел изобретательства"],
    ["Сидорова И.К.", "Трепач", "Отдел пропаганды"]
]

for row, data in enumerate(employees, start=2):
    ws.cell(row=row, column=1, value=data[0])
    ws.cell(row=row, column=2, value=data[1])
    ws.cell(row=row, column=3, value=data[2])

wb.save('../docs/employees.xlsx')


wb = load_workbook('../docs/employees.xlsx')
ws = wb.active

rows_count = ws.max_row # число заполненных строк
for row in ws.iter_rows(values_only=True):
    fio, pos, dept = row
    # print(row)
    print(f'Погоняло: {fio}, Хто: {pos}, Где: {dept}')

# работа с формулами
#
# ws['A1'] = "=СУММ(A1:A10)"
# Формат:
from openpyxl.styles import Font, Alignment

ws['A1'].font = Font(bold=True, size=14)
ws['A1'].alignment = Alignment(horizontal="center")
# etc



